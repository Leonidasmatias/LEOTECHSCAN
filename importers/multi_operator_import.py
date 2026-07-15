from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import sqlite3
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(r"C:\LEOTECHSCAN")
APP_DIR = ROOT / "APP"
BASE_DIR = ROOT / "BASE"
DATABASE_PATH = ROOT / "DATABASE" / "leotechscan.db"
LOG_PATH = ROOT / "LOGS" / "importacao_sentinel_v2.log"
EXPORT_DIR = ROOT / "EXPORTACOES"
FALLBACK_FILES = [ROOT / "BASE SPAZIO COM IBGE_n.xlsx", ROOT / "VIVO SITES.xlsx"]

TIM_COLUMNS = {
    "SITE_ID": "site_id",
    "TIPO_DE_ELEMENTO": "tipo_elemento",
    "TECNOLOGIA": "tecnologia",
    "TIPO_DE_CONEXAO": "tipo_conexao",
    "ENDERECO_ID": "endereco_id",
    "CLASSIFICACAO": "classificacao",
    "MUNICIPIO": "municipio",
    "IBGE": "ibge",
    "POPULACAO_TOTAL": "populacao",
    "ESTADO": "estado",
    "REGIONAL": "regional",
    "LATITUDE": "latitude",
    "LONGITUDE": "longitude",
    "STATUS": "status_original",
    "TIPO_DA_TORRE": "tipo_torre",
    "DETENTOR_AREA": "detentor_area",
    "DETENTOR_INFRA": "detentor_infra",
    "TIPO_DE_INFRA": "tipo_infra",
    "ALTURA_DA_ESTRUTURA": "altura",
    "SITUACAO": "situacao",
    "OTS": "ots",
    "STATION_ID": "station_id",
}

CREATE_SQL = """
CREATE TABLE sites (
 id INTEGER PRIMARY KEY,
 site TEXT,
 operadora_origem TEXT,
 municipio TEXT,
 uf TEXT,
 regional TEXT,
 latitude REAL,
 longitude REAL,
 endereco TEXT,
 status TEXT,
 projeto TEXT,
 tecnologia TEXT,
 tipo_site TEXT,
 data_importacao TEXT,
 arquivo_origem TEXT,
 site_id TEXT,
 tipo_elemento TEXT,
 tipo_conexao TEXT,
 endereco_id TEXT,
 classificacao TEXT,
 ibge TEXT,
 populacao INTEGER,
 estado TEXT,
 status_original TEXT,
 status_normalizado TEXT,
 tipo_torre TEXT,
 detentor_area TEXT,
 detentor_infra TEXT,
 tipo_infra TEXT,
 altura REAL,
 situacao TEXT,
 ots TEXT,
 geo_score INTEGER,
 risco TEXT,
 station_id TEXT,
 operadora_classificada TEXT,
 operator_rule TEXT,
 ori_score INTEGER,
 ori_risk TEXT
);
CREATE TABLE metadata (key TEXT PRIMARY KEY, value TEXT NOT NULL);
CREATE TABLE import_audit (
 id INTEGER PRIMARY KEY,
 arquivo_origem TEXT NOT NULL,
 operadora TEXT NOT NULL,
 sheet TEXT NOT NULL,
 linhas_importadas INTEGER NOT NULL,
 colunas_mapeadas INTEGER NOT NULL,
 campos_ausentes TEXT NOT NULL,
 fallback_usado INTEGER NOT NULL,
 sha256_antes TEXT NOT NULL,
 sha256_depois TEXT NOT NULL,
 excel_inalterado INTEGER NOT NULL,
 importado_em TEXT NOT NULL
);
"""

INSERT_SQL = """
INSERT INTO sites (
 id,site,operadora_origem,municipio,uf,regional,latitude,longitude,endereco,status,projeto,
 tecnologia,tipo_site,data_importacao,arquivo_origem,site_id,tipo_elemento,tipo_conexao,
 endereco_id,classificacao,ibge,populacao,estado,status_original,status_normalizado,tipo_torre,
 detentor_area,detentor_infra,tipo_infra,altura,situacao,ots,geo_score,risco,station_id,
 operadora_classificada,operator_rule,ori_score,ori_risk
) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
"""

EXPORT_COLUMNS = [
    "id",
    "site",
    "operadora_origem",
    "municipio",
    "uf",
    "regional",
    "latitude",
    "longitude",
    "endereco",
    "status",
    "projeto",
    "tecnologia",
    "tipo_site",
    "data_importacao",
    "arquivo_origem",
]


def clean(value: Any, default: str = "Nao informado") -> str:
    if value is None or value == "":
        return default
    text = str(value).strip()
    return text or default


def optional(value: Any) -> str:
    return "" if value is None else str(value).strip()


def fold(value: Any) -> str:
    text = unicodedata.normalize("NFD", clean(value, ""))
    return "".join(char for char in text.lower() if unicodedata.category(char) != "Mn")


def number(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return default


def fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_status(value: Any) -> str:
    key = fold(value)
    if key == "em aquisicao":
        return "Em aquisicao"
    if key in {"adquirido", "aquisitado"}:
        return "Aquisitado"
    if key == "desativado":
        return "Desativado"
    if key == "cancelado":
        return "Cancelado"
    return clean(value)


def geo_score(status: Any, technology: Any, population: Any, height: Any, infra: Any) -> int:
    score = 0
    status_key = fold(status)
    if "cancel" in status_key:
        score += 30
    elif "desativ" in status_key:
        score += 26
    elif "aquisicao" in status_key:
        score += 16
    elif "aquisitado" in status_key:
        score += 6
    else:
        score += 20
    tech = fold(technology)
    if any(item in tech for item in ("gsm", "tdma")):
        score += 20
    elif "umts" in tech:
        score += 16
    elif any(item in tech for item in ("lte", "multi")):
        score += 9
    elif "nr" in tech or "5g" in tech:
        score += 4
    else:
        score += 12
    pop = int(number(population))
    if pop >= 1_000_000:
        score += 20
    elif pop >= 300_000:
        score += 15
    elif pop >= 100_000:
        score += 10
    elif pop >= 30_000:
        score += 6
    else:
        score += 3
    hgt = number(height)
    if hgt >= 80:
        score += 15
    elif hgt >= 50:
        score += 12
    elif hgt >= 30:
        score += 8
    elif hgt > 0:
        score += 4
    else:
        score += 9
    infra_key = fold(infra)
    if "greenfield" in infra_key or "torre" in infra_key:
        score += 15
    elif "rooftop" in infra_key:
        score += 12
    elif "indoor" in infra_key or "smallcell" in infra_key:
        score += 5
    else:
        score += 9
    return min(100, int(score))


def risk(value: int) -> str:
    if value <= 30:
        return "Baixo"
    if value <= 60:
        return "Medio"
    if value <= 80:
        return "Alto"
    return "Critico"


def ori_score(geo: int, status: Any, height: Any, infra: Any) -> int:
    status_key = fold(status)
    if "cancel" in status_key:
        status_part = 100
    elif "desativ" in status_key:
        status_part = 90
    elif "aquisi" in status_key and "aquisitado" not in status_key:
        status_part = 60
    elif "aquisitado" in status_key:
        status_part = 20
    else:
        status_part = 65
    hgt = number(height)
    height_part = 100 if hgt >= 80 else 80 if hgt >= 50 else 55 if hgt >= 30 else 30 if hgt > 0 else 60
    infra_key = fold(infra)
    infra_part = 90 if "greenfield" in infra_key or "torre" in infra_key else 75 if "rooftop" in infra_key else 30 if "indoor" in infra_key or "smallcell" in infra_key else 60
    return max(0, min(100, round(geo * 0.4 + status_part * 0.2 + height_part * 0.2 + infra_part * 0.2)))


def first(row: dict[str, Any], *columns: str, default: str = "Nao informado") -> str:
    for column in columns:
        value = row.get(column)
        if value not in (None, ""):
            return clean(value, default)
    return default


def join_address(*parts: Any) -> str:
    return ", ".join(part for part in (optional(value) for value in parts) if part) or "Nao informado"


def detect_operator(path: Path, headers: set[str]) -> str:
    name = path.name.upper()
    if "VIVO" in name or "PMO_SIGLA" in headers or "SCIENCE_ENDERECO" in headers:
        return "VIVO"
    if "TIM" in name or "SITE_ID" in headers:
        return "TIM"
    if "CLARO" in name:
        return "CLARO"
    if "ALGAR" in name:
        return "ALGAR"
    return "NAO_IDENTIFICADO"


def find_sources(base_dir: Path, explicit_sources: list[str] | None = None) -> tuple[list[Path], bool]:
    if explicit_sources:
        return [Path(item).resolve() for item in explicit_sources], False
    base_files = sorted(path for path in base_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    if base_files:
        return base_files, False
    fallback = [path for path in FALLBACK_FILES if path.exists()]
    return fallback, True


def mapped_tim(row: dict[str, Any], imported_at: str, file_name: str, row_id: int) -> tuple[Any, ...]:
    status = normalize_status(row.get("STATUS"))
    population = int(number(row.get("POPULACAO_TOTAL")))
    height = number(row.get("ALTURA_DA_ESTRUTURA"))
    infra = clean(row.get("TIPO_DE_INFRA"))
    geo = geo_score(status, row.get("TECNOLOGIA"), population, height, infra)
    ori = ori_score(geo, status, height, infra)
    site = clean(row.get("SITE_ID"))
    address = join_address(row.get("TIPO_DE_LOGRADOURO"), row.get("LOGRADOURO"), row.get("NUMERO"), row.get("COMPLEMENTO"), row.get("BAIRRO"), row.get("CEP"))
    return (
        row_id, site, "TIM", clean(row.get("MUNICIPIO")), clean(row.get("ESTADO")), clean(row.get("REGIONAL")),
        number(row.get("LATITUDE")), number(row.get("LONGITUDE")), address, status, clean(row.get("CLASSIFICACAO")),
        clean(row.get("TECNOLOGIA")), clean(row.get("TIPO_DE_ELEMENTO")), imported_at, file_name, site,
        clean(row.get("TIPO_DE_ELEMENTO")), clean(row.get("TIPO_DE_CONEXAO")), clean(row.get("ENDERECO_ID")),
        clean(row.get("CLASSIFICACAO")), clean(row.get("IBGE")), population, clean(row.get("ESTADO")),
        clean(row.get("STATUS")), status, clean(row.get("TIPO_DA_TORRE")), clean(row.get("DETENTOR_AREA")),
        clean(row.get("DETENTOR_INFRA")), infra, height, clean(row.get("SITUACAO")), clean(row.get("OTS")),
        geo, risk(geo), clean(row.get("STATION_ID")), "TIM", "SOURCE_FILE:TIM", ori, risk(ori)
    )


def mapped_vivo(row: dict[str, Any], imported_at: str, file_name: str, row_id: int) -> tuple[Any, ...]:
    site = first(row, "PMO_SIGLA", "SIGLA_LOGICA_REFERENCIA", "UID_IDMASTER", "UID_IDPMTS")
    status = first(row, "MASTEROBRA_STATUS_ROLLOUT", "VALID_STATUS_ROLLOUT", "TA_STATUS", "FAROL_ESTEIRA")
    technology = first(row, "TECNOLOGIAS_FINAIS", "PMO_TECN_EQUIP", "PMO_FREQ_EQUIP")
    project = first(row, "PMO_PROJETO", "MASTEROBRA_MEGA_PROJETO", "PMO_SUB_PROJETO", "PMO_TIPO_OBRA")
    tipo_site = first(row, "PMO_TIPO_OBRA", "PMO_TIPO_PMTS", "ENG_ESTRUTURA_TIPO", "MASTERSITE_COBERTURA")
    infra = first(row, "ENG_ESTRUTURA_TIPO", "MASTERSITE_COBERTURA", "PMO_TIPO_OBRA")
    address = join_address(row.get("SCIENCE_ENDERECO"), row.get("SCIENCE_COMPLEMENTO"), row.get("SCIENCE_BAIRRO"), row.get("SCIENCE_CEP"))
    geo = geo_score(status, technology, 0, 0, infra)
    ori = ori_score(geo, status, 0, infra)
    uf = first(row, "PMO_UF", "UID_UFSIGLA")
    return (
        row_id, site, "VIVO", first(row, "IBGE_MUNICIPIO"), uf, first(row, "PMO_REGIONAL", "REGIONAL_CARIMBO"),
        number(row.get("SCIENCE_LATITUDE")), number(row.get("SCIENCE_LONGITUDE")), address, status, project,
        technology, tipo_site, imported_at, file_name, site, tipo_site, "Nao informado", first(row, "UID_IDENG", "UID_IDPMTS"),
        project, clean(row.get("UID_IBGE")), 0, uf, status, status, tipo_site, first(row, "RSO_RSA_DETENTORA", "SCIENCE_DETENTORA"),
        first(row, "RSO_RSA_DETENTORA", "SCIENCE_DETENTORA"), infra, 0.0, first(row, "FAROL_ESTEIRA"), first(row, "TX_STATUS_ACIONAMENTO"),
        geo, risk(geo), first(row, "UID_IDMASTER", "UID_IDPMTS"), "VIVO", "SOURCE_FILE:VIVO", ori, risk(ori)
    )


def read_sheet(path: Path, fallback_used: bool, conn: sqlite3.Connection, start_id: int, imported_at: str) -> tuple[int, dict[str, Any]]:
    before_hash = fingerprint(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value, "") for value in next(rows)]
    indexes = {header: index for index, header in enumerate(headers) if header}
    operator = detect_operator(path, set(indexes))
    missing: list[str] = []
    if operator == "TIM":
        missing = [column for column in TIM_COLUMNS if column not in indexes]
    elif operator == "VIVO":
        required_any = ["PMO_SIGLA", "IBGE_MUNICIPIO", "PMO_UF"]
        missing = [column for column in required_any if column not in indexes]
    else:
        logging.warning("file_skipped_unknown_operator file=%s", path.name)
        workbook.close()
        return start_id, {"file": path.name, "operator": operator, "rows": 0, "skipped": True}
    if operator == "TIM" and missing:
        raise RuntimeError("Colunas TIM obrigatorias ausentes em %s: %s" % (path.name, ", ".join(missing)))

    row_id = start_id
    batch: list[tuple[Any, ...]] = []
    imported = 0
    mapper = mapped_vivo if operator == "VIVO" else mapped_tim
    for raw in rows:
        if not any(value is not None for value in raw):
            continue
        row = {header: raw[index] if index < len(raw) else None for header, index in indexes.items()}
        row_id += 1
        imported += 1
        batch.append(mapper(row, imported_at, path.name, row_id))
        if len(batch) >= 5000:
            conn.executemany(INSERT_SQL, batch)
            conn.commit()
            batch.clear()
    if batch:
        conn.executemany(INSERT_SQL, batch)
        conn.commit()
    workbook.close()
    after_hash = fingerprint(path)
    conn.execute(
        "INSERT INTO import_audit(arquivo_origem,operadora,sheet,linhas_importadas,colunas_mapeadas,campos_ausentes,fallback_usado,sha256_antes,sha256_depois,excel_inalterado,importado_em) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (path.name, operator, sheet.title, imported, len(indexes), json.dumps(missing, ensure_ascii=False), 1 if fallback_used else 0, before_hash, after_hash, 1 if before_hash == after_hash else 0, imported_at),
    )
    conn.commit()
    logging.info("file_imported file=%s operator=%s rows=%d fallback=%s unchanged=%s", path.name, operator, imported, fallback_used, before_hash == after_hash)
    return row_id, {"file": path.name, "operator": operator, "rows": imported, "sha256": before_hash, "unchanged": before_hash == after_hash}


def write_csv(path: Path, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream, delimiter=";", quoting=csv.QUOTE_ALL)
        writer.writerow(headers)
        writer.writerows(rows)


def export_reports(conn: sqlite3.Connection, export_dir: Path) -> None:
    export_dir.mkdir(parents=True, exist_ok=True)
    write_csv(
        export_dir / "auditoria_importacao.csv",
        ["arquivo_origem", "operadora", "sheet", "linhas_importadas", "fallback_usado", "excel_inalterado", "importado_em"],
        conn.execute("SELECT arquivo_origem,operadora,sheet,linhas_importadas,fallback_usado,excel_inalterado,importado_em FROM import_audit ORDER BY id"),
    )
    write_csv(
        export_dir / "sites_consolidados.csv",
        EXPORT_COLUMNS,
        conn.execute("SELECT %s FROM sites ORDER BY operadora_origem,site" % ",".join(EXPORT_COLUMNS)),
    )
    write_csv(
        export_dir / "sites_por_operadora.csv",
        ["operadora_origem", "registros", "sites_unicos", "ufs", "municipios"],
        conn.execute("SELECT operadora_origem,COUNT(*),COUNT(DISTINCT site),COUNT(DISTINCT uf),COUNT(DISTINCT municipio) FROM sites GROUP BY operadora_origem ORDER BY COUNT(*) DESC"),
    )


def import_all(sources: list[Path], database: Path, log_path: Path, export_dir: Path, fallback_used: bool) -> dict[str, Any]:
    database.parent.mkdir(parents=True, exist_ok=True)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    export_dir.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(filename=log_path, level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", encoding="utf-8")
    if fallback_used:
        logging.warning("base_folder_empty_fallback_used files=%s", ",".join(path.name for path in sources))
    temp = database.with_suffix(".tmp.db")
    if temp.exists():
        temp.unlink()
    conn = sqlite3.connect(temp)
    conn.executescript(CREATE_SQL)
    imported_at = datetime.now(timezone.utc).isoformat()
    row_id = 0
    summaries = []
    for source in sources:
        if source.name.startswith("~$") or source.suffix.lower() != ".xlsx":
            continue
        row_id, summary = read_sheet(source, fallback_used, conn, row_id, imported_at)
        summaries.append(summary)
    conn.executemany(
        "INSERT INTO metadata(key,value) VALUES (?,?)",
        {
            "schema_version": "2.0-sprint1",
            "imported_at": imported_at,
            "source_name": ", ".join(path.name for path in sources),
            "row_count": str(row_id),
            "fallback_used": str(fallback_used),
            "operators": json.dumps({row[0]: row[1] for row in conn.execute("SELECT operadora_origem,COUNT(*) FROM sites GROUP BY operadora_origem")}, ensure_ascii=False),
        }.items(),
    )
    conn.executescript(
        "CREATE INDEX idx_filters ON sites(estado,municipio,tecnologia,status_normalizado,detentor_infra,tipo_infra,operadora_classificada);"
        "CREATE INDEX idx_score ON sites(geo_score DESC);"
        "CREATE INDEX idx_site ON sites(site_id);"
        "CREATE INDEX idx_unified_site ON sites(site,operadora_origem,uf);"
    )
    export_reports(conn, export_dir)
    conn.commit()
    conn.execute("PRAGMA optimize")
    total = conn.execute("SELECT COUNT(*) FROM sites").fetchone()[0]
    distribution = dict(conn.execute("SELECT operadora_origem,COUNT(*) FROM sites GROUP BY operadora_origem"))
    conn.close()
    temp.replace(database)
    logging.info("import_finished rows=%d database=%s distribution=%s", total, database, json.dumps(distribution, ensure_ascii=False))
    return {"rows": total, "database": str(database), "fallback_used": fallback_used, "sources": summaries, "operators": distribution}


def main() -> None:
    parser = argparse.ArgumentParser(description="Sentinel-1 V2 Sprint 1 multi-operator importer")
    parser.add_argument("--base-dir", default=str(BASE_DIR))
    parser.add_argument("--database", default=str(DATABASE_PATH))
    parser.add_argument("--log", default=str(LOG_PATH))
    parser.add_argument("--export-dir", default=str(EXPORT_DIR))
    parser.add_argument("--source", action="append", help="Arquivo XLSX explicito. Pode ser usado mais de uma vez.")
    args = parser.parse_args()
    sources, fallback_used = find_sources(Path(args.base_dir), args.source)
    if not sources:
        raise SystemExit("Nenhum arquivo .xlsx encontrado em BASE e nenhum fallback disponivel.")
    result = import_all([path.resolve() for path in sources], Path(args.database).resolve(), Path(args.log).resolve(), Path(args.export_dir).resolve(), fallback_used)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        logging.exception("import_failed type=%s", type(exc).__name__)
        raise
