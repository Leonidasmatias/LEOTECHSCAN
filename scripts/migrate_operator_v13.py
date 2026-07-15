from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
import shutil
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def classify(site_id: object, station_id: object, technology: object, rules: dict) -> tuple[str, str]:
    site, station, tech = clean(site_id).upper(), clean(station_id).upper(), clean(technology).upper()
    for rule in rules["operators"]:
        name = rule["name"]
        for prefix in rule.get("site_prefixes", []):
            if site.startswith(prefix.upper()): return name, f"SITE_PREFIX:{prefix}"
        for prefix in rule.get("station_prefixes", []):
            if station.startswith(prefix.upper()): return name, f"STATION_PREFIX:{prefix}"
        for marker in rule.get("contains", []):
            if marker.upper() in site or marker.upper() in station: return name, f"MARKER:{marker}"
        for pattern in rule.get("technology_patterns", []):
            if re.search(pattern, tech, flags=re.IGNORECASE): return name, f"TECH:{pattern}"
    return "Não Identificado", "NO_RULE_MATCH"


def status_score(status: object) -> int:
    value = clean(status).lower()
    if "cancel" in value: return 100
    if "desativ" in value: return 90
    if "aquisi" in value and "aquisitado" not in value: return 60
    if "aquisitado" in value: return 20
    return 65


def height_score(height: object) -> int:
    try: value = float(height or 0)
    except (TypeError, ValueError): value = 0
    if value >= 80: return 100
    if value >= 50: return 80
    if value >= 30: return 55
    if value > 0: return 30
    return 60


def infra_score(infra: object) -> int:
    value = clean(infra).lower()
    if "greenfield" in value or "torre" in value: return 90
    if "rooftop" in value: return 75
    if "indoor" in value or "smallcell" in value: return 30
    return 60


def ori(geo: object, status: object, height: object, infra: object) -> int:
    value = round(float(geo or 0) * .4 + status_score(status) * .2 + height_score(height) * .2 + infra_score(infra) * .2)
    return max(0, min(100, value))


def risk(value: int) -> str:
    if value <= 30: return "Baixo"
    if value <= 60: return "Médio"
    if value <= 80: return "Alto"
    return "Crítico"


def main() -> None:
    parser = argparse.ArgumentParser(description="Migração atômica LeoTechScan V1.3 Operator Intelligence")
    parser.add_argument("--source", required=True)
    parser.add_argument("--database", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--rules", required=True)
    parser.add_argument("--log", required=True)
    args = parser.parse_args()
    source, database, output, rule_path = map(lambda p: Path(p).resolve(), (args.source, args.database, args.output, args.rules))
    log_path = Path(args.log).resolve(); output.parent.mkdir(parents=True, exist_ok=True); log_path.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(filename=log_path, level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", encoding="utf-8")
    rules = json.loads(rule_path.read_text(encoding="utf-8"))
    rules_hash = hashlib.sha256(rule_path.read_bytes()).hexdigest()
    if output.exists(): output.unlink()
    shutil.copy2(database, output)
    logging.info("v13_migration_started source=%s database=%s rules_version=%s", source.name, database.name, rules["version"])
    wb = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    ws = wb[wb.sheetnames[0]]; rows = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]; index = {name: position for position, name in enumerate(headers)}
    required = ["SITE_ID", "STATION_ID", "TECNOLOGIA"]
    missing = [name for name in required if name not in index]
    if missing: raise RuntimeError("Colunas ausentes: " + ", ".join(missing))
    conn = sqlite3.connect(output)
    existing = {row[1] for row in conn.execute("PRAGMA table_info(sites)")}
    additions = {"station_id": "TEXT", "operadora_classificada": "TEXT", "operator_rule": "TEXT", "ori_score": "INTEGER", "ori_risk": "TEXT"}
    for column, kind in additions.items():
        if column not in existing: conn.execute(f"ALTER TABLE sites ADD COLUMN {column} {kind}")
    select = conn.execute("SELECT id,site_id,tecnologia,status_normalizado,altura,tipo_infra,geo_score FROM sites ORDER BY id")
    batch, totals, processed = [], {}, 0
    data_rows = (row for row in rows if any(value is not None for value in row))
    for excel_row, db_row in zip(data_rows, select):
        row_id, db_site, db_tech, status, height, infra, geo = db_row
        excel_site = clean(excel_row[index["SITE_ID"]]); station = clean(excel_row[index["STATION_ID"]]); excel_tech = clean(excel_row[index["TECNOLOGIA"]])
        if (excel_site or "Não informado") != clean(db_site) or (excel_tech or "Não informado") != clean(db_tech):
            raise RuntimeError(f"Divergência de alinhamento na linha {row_id}: excel=({excel_site!r},{excel_tech!r}) db=({clean(db_site)!r},{clean(db_tech)!r})")
        operator, matched_rule = classify(excel_site, station, excel_tech, rules); score = ori(geo, status, height, infra)
        batch.append((station or "Não informado", operator, matched_rule, score, risk(score), row_id)); totals[operator] = totals.get(operator, 0) + 1; processed += 1
        if len(batch) >= 5000:
            conn.executemany("UPDATE sites SET station_id=?,operadora_classificada=?,operator_rule=?,ori_score=?,ori_risk=? WHERE id=?", batch); conn.commit(); batch.clear()
    if batch: conn.executemany("UPDATE sites SET station_id=?,operadora_classificada=?,operator_rule=?,ori_score=?,ori_risk=? WHERE id=?", batch)
    if processed != conn.execute("SELECT COUNT(*) FROM sites").fetchone()[0]: raise RuntimeError("Quantidade migrada diverge do SQLite")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_operator ON sites(operadora_classificada)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ori ON sites(ori_score DESC)")
    conn.execute("INSERT OR REPLACE INTO metadata(key,value) VALUES ('schema_version','1.3'),('operator_rules_version',?),('operator_rules_sha256',?)", (rules["version"], rules_hash))
    conn.commit(); conn.execute("PRAGMA optimize"); conn.close(); wb.close()
    logging.info("v13_migration_finished rows=%d distribution=%s rules_sha256_prefix=%s", processed, json.dumps(totals, ensure_ascii=False), rules_hash[:12])
    print(json.dumps({"rows": processed, "operators": totals, "output": str(output), "rules_sha256": rules_hash}, ensure_ascii=False))


if __name__ == "__main__":
    main()
